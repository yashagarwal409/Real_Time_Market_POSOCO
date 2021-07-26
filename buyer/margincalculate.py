from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def calculate():
    dec = load_workbook(r"C:\Users\yagar\Downloads\dec.xlsx")
    sec = load_workbook(r"C:\Users\yagar\Downloads\sec.xlsx")
    mar = Workbook()
    mars = mar.active
    decs = dec.active
    secs = sec.active
    for row in range(5, 102):
        for col in range(1, 15):
            cell = get_column_letter(col)
            if(row == 5 or col <= 2):
                mars[cell+str(row)] = decs[cell+str(row)].value
            else:
                mars[cell+str(row)] = decs[cell+str(row)].value + \
                    secs[cell+str(row)].value
    mar.save(r"C:\Users\yagar\Downloads\margin.xlsx")
