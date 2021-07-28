import datetime
#from datetime import datetime, time
from django.core.checks import messages
from django.shortcuts import render, redirect, HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from .models import Declaration, Reserve, Schedule, TimeBlock
from .extractdata import extract
from .margincalculate import calculate
from .extractdatadat import extractdat
from .margincalculatedat import calculatedat
from .forms import UpReserveForm, rtmform
from seller.models import Bid
from django.conf import settings
from django.urls import reverse
User = settings.AUTH_USER_MODEL


def is_time_between(begin_time, end_time):
    # If check time is not given, default to current UTC time
    check_time = timezone.localtime(timezone.now()).time()
    if begin_time < end_time:
        return check_time >= begin_time and check_time <= end_time
    else:  # crosses midnight
        return check_time >= begin_time or check_time <= end_time


def home(response):
    return render(response, 'buyer/home.html', {})


def update(response):
    return render(response, "buyer/update.html", {})


def display(response):
    return render(response, 'buyer/display.html', {})


def rtm(response):
    return render(response, 'buyer/rtm.html', {})


def dat(response):
    return render(response, 'buyer/dat.html', {})


def updatertm(response):
    mar = load_workbook(r'C:\Users\yagar\Downloads\margin.xlsx')
    Reserve.objects.filter(date=timezone.now()).delete()
    margin = mar.active
    for row in range(6, 102):
        for col in range(3, 15):
            cell = get_column_letter(col)
            name = str(margin[cell+'5'].value).partition('(')[0]
            if(name == 'AGBPP-GAS' or name == 'AGTCCPP-GAS' or name == 'BGTPP' or name == 'PALATANA-GAS'):
                Reserve.objects.create(
                    time=margin['B'+str(row)].value, name=name, quantity=margin[cell+str(row)].value)
    mar = load_workbook(r'C:\Users\yagar\Downloads\dec.xlsx')
    Declaration.objects.filter(date=timezone.now()).delete()
    margin = mar.active
    for row in range(6, 102):
        for col in range(3, 15):
            cell = get_column_letter(col)
            name = str(margin[cell+'5'].value).partition('(')[0]
            if(name == 'AGBPP-GAS' or name == 'AGTCCPP-GAS' or name == 'BGTPP' or name == 'PALATANA-GAS'):
                Declaration.objects.create(
                    time=margin['B'+str(row)].value, name=name, quantity=margin[cell+str(row)].value)
    mar = load_workbook(r'C:\Users\yagar\Downloads\sec.xlsx')
    Schedule.objects.filter(date=timezone.now()).delete()
    margin = mar.active
    for row in range(6, 102):
        for col in range(3, 15):
            cell = get_column_letter(col)
            name = str(margin[cell+'5'].value).partition('(')[0]
            if(name == 'AGBPP-GAS' or name == 'AGTCCPP-GAS' or name == 'BGTPP' or name == 'PALATANA-GAS'):
                Schedule.objects.create(
                    time=margin['B'+str(row)].value, name=name, quantity=margin[cell+str(row)].value)
    return render(response, 'buyer/message.html', {"message": "The Real Time Databases Have Been Updated Successfully!"})


def updatedat(response):
    mar = load_workbook(r'C:\Users\yagar\Downloads\margin1.xlsx')
    date = timezone.now() + datetime.timedelta(days=1)
    Reserve.objects.filter(date=date).delete()
    margin = mar.active
    for row in range(6, 102):
        for col in range(3, 15):
            cell = get_column_letter(col)
            name = str(margin[cell+'5'].value).partition('(')[0]
            if(name == 'AGBPP-GAS' or name == 'AGTCCPP-GAS' or name == 'BGTPP' or name == 'PALATANA-GAS'):
                Reserve.objects.create(date=date,
                                       time=margin['B'+str(row)].value, name=name, quantity=margin[cell+str(row)].value)
    mar = load_workbook(r'C:\Users\yagar\Downloads\dec1.xlsx')
    Declaration.objects.filter(date=date).delete()
    margin = mar.active
    for row in range(6, 102):
        for col in range(3, 15):
            cell = get_column_letter(col)
            name = str(margin[cell+'5'].value).partition('(')[0]
            if(name == 'AGBPP-GAS' or name == 'AGTCCPP-GAS' or name == 'BGTPP' or name == 'PALATANA-GAS'):
                Declaration.objects.create(date=date,
                                           time=margin['B'+str(row)].value, name=name, quantity=margin[cell+str(row)].value)
    mar = load_workbook(r'C:\Users\yagar\Downloads\sec1.xlsx')
    Schedule.objects.filter(date=date).delete()
    margin = mar.active
    for row in range(6, 102):
        for col in range(3, 15):
            cell = get_column_letter(col)
            name = str(margin[cell+'5'].value).partition('(')[0]
            if(name == 'AGBPP-GAS' or name == 'AGTCCPP-GAS' or name == 'BGTPP' or name == 'PALATANA-GAS'):
                Schedule.objects.create(date=date,
                                        time=margin['B'+str(row)].value, name=name, quantity=margin[cell+str(row)].value)
    return render(response, 'buyer/message.html', {"message": "The Day Ahead Databases Have Been Updated Successfully!"})


def refreshrtm(response):
    extract()
    calculate()
    return render(response, 'buyer/message.html', {"message": "The New Real Time Data Has Been Downloaded Successfully!"})


def refreshdat(response):
    extractdat()
    calculatedat()
    return render(response, 'buyer/message.html', {"message": "The New Real Time Data Has Been Downloaded Successfully!"})


def reserve(response, datorrtm):
    if(datorrtm == "dat"):
        date = timezone.now() + datetime.timedelta(days=1)
    elif datorrtm == "rtm":
        date = timezone.now()
    list1 = Reserve.objects.filter(name='AGBPP-GAS', date=date)
    list2 = Reserve.objects.filter(name='AGTCCPP-GAS', date=date)
    list3 = Reserve.objects.filter(name='BGTPP', date=date)
    list4 = Reserve.objects.filter(name='PALATANA-GAS', date=date)
    list5 = []
    for i in range(0, 96):
        total = 0
        total = list1[i].quantity+list2[i].quantity + \
            list3[i].quantity+list4[i].quantity
        list5.append(total)
    return render(response, 'buyer/displaydata.html', {"list1": list1, "list2": list2, "list3": list3, "list4": list4, "list5": list5})


def declaration(response, datorrtm):
    if(datorrtm == "dat"):
        date = timezone.now() + datetime.timedelta(days=1)
    elif datorrtm == "rtm":
        date = timezone.now()
    list1 = Declaration.objects.filter(name='AGBPP-GAS', date=date)
    list2 = Declaration.objects.filter(name='AGTCCPP-GAS', date=date)
    list3 = Declaration.objects.filter(name='BGTPP', date=date)
    list4 = Declaration.objects.filter(name='PALATANA-GAS', date=date)
    list5 = []
    for i in range(0, 96):
        total = 0
        total = list1[i].quantity+list2[i].quantity + \
            list3[i].quantity+list4[i].quantity
        list5.append(total)
    return render(response, 'buyer/displaydata.html', {"list1": list1, "list2": list2, "list3": list3, "list4": list4, "list5": list5})


def schedule(response, datorrtm):
    if(datorrtm == "dat"):
        date = timezone.now() + datetime.timedelta(days=1)
    elif datorrtm == "rtm":
        date = timezone.now()
    list1 = Schedule.objects.filter(name='AGBPP-GAS', date=date)
    list2 = Schedule.objects.filter(name='AGTCCPP-GAS', date=date)
    list3 = Schedule.objects.filter(name='BGTPP', date=date)
    list4 = Schedule.objects.filter(name='PALATANA-GAS', date=date)
    list5 = []
    for i in range(0, 96):
        total = 0
        total = list1[i].quantity+list2[i].quantity + \
            list3[i].quantity+list4[i].quantity
        list5.append(total)
    return render(response, 'buyer/displaydata.html', {"list1": list1, "list2": list2, "list3": list3, "list4": list4, "list5": list5})


def upreserve(response):
    if response.method == 'POST':
        form = UpReserveForm(response.POST)
        user = response.user
        if(form.is_valid()):
            bids = Bid.objects.filter(
                time=form.cleaned_data['time'], date=form.cleaned_data['date'], is_up=True, is_dat=True).order_by('price')
            quantity = float(form.cleaned_data['quantity'])
            sum = 0
            for bid in bids:
                sum = sum+float(bid.volume)
                print(bid.seller.username, bid.volume)
            if sum < quantity:
                return HttpResponse("The entered amount cannot be cleared.")
            else:
                mcp = 0
                for bid in bids:
                    if(float(bid.volume) < float(quantity)):
                        quantity = quantity-float(bid.volume)
                    else:
                        mcp = float(bid.price)
                        print(mcp)
                        break
                quantity = float(form.cleaned_data['quantity'])
                cleared = user.clearedreserveup_set.create(
                    mcp=mcp, time_block=form.cleaned_data['time'], date=form.cleaned_data['date'], is_dat=True)
                for bid in bids:
                    if(float(bid.volume) < float(quantity)):
                        cleared.clearentityup_set.create(
                            name=bid.seller.username, amount=bid.volume)
                        quantity = quantity-float(bid.volume)
                    else:
                        cleared.clearentityup_set.create(
                            name=bid.seller.username, amount=quantity)
                        break
            # user.upreserve_set.create(time=form.cleaned_data['time'], quantity=form.cleaned_data['quantity'])
    else:
        form = UpReserveForm()
    return render(response, 'buyer/upreserve.html', {"form": form})


def downreserve(response):
    if response.method == 'POST':
        form = UpReserveForm(response.POST)
        user = response.user
        if(form.is_valid()):
            bids = Bid.objects.filter(
                time=form.cleaned_data['time'], date=form.cleaned_data['date'], is_down=True, is_dat=True).order_by('-price')
            quantity = float(form.cleaned_data['quantity'])
            # sum = 0
            # for bid in bids:
            #  sum = sum+float(bid.volume)
            #  print(bid.seller.username, bid.volume)
            # if sum < quantity:
            #   return HttpResponse("The entered amount cannot be cleared.")
            quantity = float(form.cleaned_data['quantity'])
            cleared = user.clearedreservedown_set.create(
                time_block=form.cleaned_data['time'], date=form.cleaned_data['date'], is_dat=True)
            for bid in bids:
                if(float(bid.volume) < float(quantity)):
                    cleared.clearentitydown_set.create(
                        name=bid.seller.username, amount=bid.volume, price=bid.price)
                    quantity = quantity-float(bid.volume)
                else:
                    cleared.clearentitydown_set.create(
                        name=bid.seller.username, amount=quantity, price=bid.price)
                    break
            # user.upreserve_set.create(time=form.cleaned_data['time'], quantity=form.cleaned_data['quantity'])
    else:
        form = UpReserveForm()
    return render(response, 'buyer/upreserve.html', {"form": form})


def upreservertm(response):
    is_rtm = False
    is_up = False
    is_dat = False
    is_down = False
    form = None
    if response.method == 'POST':
        user = response.user
        form = response.POST
        messagelist = []
        bids = Bid.objects.filter(
            time=form.get('time'), date=form.get('date'), is_up=True, is_rtm=True).order_by('price')
        quantity = float(form.get('quantity'))
        sum = 0
        time = form.get('time')
        for bid in bids:
            sum = sum+float(bid.volume)
        if sum == 0 or quantity == 0:
            message = "Time Block:"+time+" Cleared Quantity: 0"
            messagelist.append(message)
            return render(response, 'buyer/success.html', {"messages": messagelist})
        if sum < quantity:
            message = "Time Block:"+time+" Cleared Quantity: "+str(sum)
            messagelist.append(message)
            # return redirect(reverse('runcodedatup'))
            # return HttpResponse("The entered amount cannot be cleared.")
        else:
            message = "Time Block:"+time + \
                " Cleared Quantity: "+str(quantity)
            messagelist.append(message)
        for bid in bids:
            mcp = 0
            for bid in bids:
                mcp = float(bid.price)
            for bid in bids:
                if(float(bid.volume) < float(quantity)):
                    quantity = quantity-float(bid.volume)
                else:
                    mcp = float(bid.price)
                    break
        quantity = float(form.get('quantity'))
        cleared = user.clearedreserveup_set.create(
            mcp=mcp, time_block=form.get('time'), date=form.get('date'), is_rtm=True)
        for bid in bids:
            if(float(bid.volume) < float(quantity)):
                cleared.clearentityup_set.create(
                    name=bid.seller.username, amount=bid.volume)
                quantity = quantity-float(bid.volume)
            else:
                cleared.clearentityup_set.create(
                    name=bid.seller.username, amount=quantity)
                break
        return render(response, 'buyer/success.html', {"messages": messagelist})
    objectlist = TimeBlock.objects.all()
    timelist = []
    for object in objectlist:
        timelist.append(object.time)
    for i in range(1, 96, 2):
        times = timelist[i]
        begt = datetime.datetime.strptime(times[0:5], '%H:%M').time()
        if i != 95:
            endt = datetime.datetime.strptime(times[6:11], '%H:%M').time()
        else:
            endt = datetime.datetime.strptime('00:00', '%H:%M').time()
        if(is_time_between(begt, endt)):
            form = rtmform(
                timeoptions=[timelist[(i+5) % 96], timelist[(i+6) % 96]])
            break
    if form == None:
        return render(response, 'buyer/message.html', {"message": "Please login after a while. The portal is down."})
    return render(response, 'buyer/upreserve.html', {'form': form})


def downreservertm(response):
    is_rtm = False
    is_up = False
    is_dat = False
    is_down = False
    form = None
    if response.method == 'POST':
        user = response.user
        form = response.POST
        messagelist = []
        bids = Bid.objects.filter(
            time=form.get('time'), date=form.get('date'), is_down=True, is_rtm=True).order_by('-price')
        quantity = float(form.get('quantity'))
        sum = 0
        time = form.get('time')
        for bid in bids:
            sum = sum+float(bid.volume)
        if sum == 0 or quantity == 0:
            message = "Time Block:"+time+" Cleared Quantity: 0"
            messagelist.append(message)
            return render(response, 'buyer/success.html', {"messages": messagelist})
        if sum < quantity:
            message = "Time Block:"+time+" Cleared Quantity: "+str(sum)
            messagelist.append(message)
            # return redirect(reverse('runcodedatup'))
            # return HttpResponse("The entered amount cannot be cleared.")
        else:
            message = "Time Block:"+time + \
                " Cleared Quantity: "+str(quantity)
            messagelist.append(message)
        quantity = float(form.get('quantity'))
        cleared = user.clearedreservedown_set.create(
            time_block=form.get('time'), date=form.get('date'), is_rtm=True)
        for bid in bids:
            if(float(bid.volume) < float(quantity)):
                cleared.clearentitydown_set.create(
                    name=bid.seller.username, amount=bid.volume, price=bid.price)
                quantity = quantity-float(bid.volume)
            else:
                cleared.clearentitydown_set.create(
                    name=bid.seller.username, amount=quantity, price=bid.price)
                break
        return render(response, 'buyer/success.html', {"messages": messagelist})
    objectlist = TimeBlock.objects.all()
    timelist = []
    for object in objectlist:
        timelist.append(object.time)
    for i in range(1, 96, 2):
        times = timelist[i]
        begt = datetime.datetime.strptime(times[0:5], '%H:%M').time()
        if i != 95:
            endt = datetime.datetime.strptime(times[6:11], '%H:%M').time()
        else:
            endt = datetime.datetime.strptime('00:00', '%H:%M').time()
        if(is_time_between(begt, endt)):
            form = rtmform(
                timeoptions=[timelist[(i+5) % 96], timelist[(i+6) % 96]])
            break
    if form == None:
        return render(response, 'buyer/message.html', {"message": "Please login after a while. The portal is down."})
    return render(response, 'buyer/upreserve.html', {'form': form})


def upreservedat(response):
    is_rtm = False
    is_up = False
    is_dat = False
    is_down = False
    form = None
    if response.method == 'POST':
        user = response.user
        form = response.POST
        bids = Bid.objects.filter(
            time=form.get('time'), date=form.get('date'), is_up=True, is_dat=True).order_by('price')
        quantity = float(form.get('quantity'))
        sum = 0
        for bid in bids:
            sum = sum+float(bid.volume)
        if sum < quantity:
            return HttpResponse("The entered amount cannot be cleared.")
        else:
            mcp = 0
            for bid in bids:
                if(float(bid.volume) < float(quantity)):
                    quantity = quantity-float(bid.volume)
                else:
                    mcp = float(bid.price)
                    break
            quantity = float(form.get('quantity'))
            cleared = user.clearedreserveup_set.create(
                mcp=mcp, time_block=form.get('time'), date=form.get('date'), is_dat=True)
            for bid in bids:
                if(float(bid.volume) < float(quantity)):
                    cleared.clearentityup_set.create(
                        name=bid.seller.username, amount=bid.volume)
                    quantity = quantity-float(bid.volume)
                else:
                    cleared.clearentityup_set.create(
                        name=bid.seller.username, amount=quantity)
                    break
            return HttpResponse("The entered amount has been cleared.")
    objectlist = Reserve.objects.filter(name='AGBPP-GAS').order_by('time')
    timelist = []
    for object in objectlist:
        timelist.append(object.time)
    form = rtmform(
        timeoptions=timelist)
    return render(response, 'buyer/upreserve.html', {'form': form})


def downreservedat(response):
    is_rtm = False
    is_up = False
    is_dat = False
    is_down = False
    form = None
    if response.method == 'POST':
        user = response.user
        form = response.POST
        bids = Bid.objects.filter(
            time=form.get('time'), date=form.get('date'), is_down=True, is_dat=True).order_by('-price')
        quantity = float(form.get('quantity'))
        sum = 0
        for bid in bids:
            sum = sum+float(bid.volume)
        if sum < quantity:
            return HttpResponse("The entered amount cannot be cleared.")
        else:
            quantity = float(form.get('quantity'))
            cleared = user.clearedreservedown_set.create(
                time_block=form.get('time'), date=form.get('date'), is_dat=True)
            for bid in bids:
                if(float(bid.volume) < float(quantity)):
                    cleared.clearentitydown_set.create(
                        name=bid.seller.username, amount=bid.volume, price=bid.price)
                    quantity = quantity-float(bid.volume)
                else:
                    cleared.clearentitydown_set.create(
                        name=bid.seller.username, amount=quantity, price=bid.price)
                    break
            return HttpResponse("The entered amount has been cleared.")
    objectlist = Reserve.objects.filter(name='AGBPP-GAS').order_by('time')
    timelist = []
    for object in objectlist:
        timelist.append(object.time)
    form = rtmform(
        timeoptions=timelist)
    return render(response, 'buyer/upreserve.html', {'form': form})


def dispclearup(response, type):
    is_rtm = False
    is_dat = False
    if(type == "dat"):
        is_dat = True
    if(type == "rtm"):
        is_rtm = True
    clearedreserve = response.user.clearedreserveup_set.filter(
        is_dat=is_dat, is_rtm=is_rtm).order_by('date').order_by('time_block')
    return render(response, 'buyer/clearedupdata.html', {"clearedreserve": clearedreserve})


def dispcleardown(response, type):
    is_rtm = False
    is_dat = False
    if(type == "dat"):
        is_dat = True
    if(type == "rtm"):
        is_rtm = True
    clearedreserve = response.user.clearedreservedown_set.filter(
        is_dat=is_dat, is_rtm=is_rtm)
    return render(response, 'buyer/cleareddowndata.html', {'clearedreserve': clearedreserve})


def runcodedatup(response):
    if response.method == 'POST':
        user = response.user
        form = response.POST
        messagelist = []
        for time in TimeBlock.objects.all():
            id = 'q'+str(time.id)
            val = response.POST.get(id)
            if float(val) > 0:
                time = time.time
                bids = Bid.objects.filter(
                    time=time, date=response.POST.get('date'), is_up=True, is_dat=True).order_by('price')
                quantity = float(val)
                sum = 0
                for bid in bids:
                    sum = sum+float(bid.volume)
                if sum == 0:
                    message = "Time Block:"+time+" Cleared Quantity: 0"
                    messagelist.append(message)
                    continue
                if sum < quantity:
                    message = "Time Block:"+time+" Cleared Quantity: "+str(sum)
                    messagelist.append(message)
                    # return redirect(reverse('runcodedatup'))
                    # return HttpResponse("The entered amount cannot be cleared.")
                else:
                    message = "Time Block:"+time + \
                        " Cleared Quantity: "+str(quantity)
                    messagelist.append(message)
                for bid in bids:
                    mcp = float(bid.price)
                for bid in bids:
                    if(float(bid.volume) < float(quantity)):
                        quantity = quantity-float(bid.volume)
                    else:
                        mcp = float(bid.price)
                        break
                quantity = float(val)
                cleared = user.clearedreserveup_set.create(
                    mcp=mcp, time_block=time, date=response.POST.get('date'), is_dat=True)
                for bid in bids:
                    if(float(bid.volume) < float(quantity)):
                        cleared.clearentityup_set.create(
                            name=bid.seller.username, amount=bid.volume)
                        quantity = quantity-float(bid.volume)
                    else:
                        cleared.clearentityup_set.create(
                            name=bid.seller.username, amount=quantity)
                        break
        return render(response, 'buyer/success.html', {"messages": messagelist})
    else:
        return render(response, "buyer/datform.html", {'timeblock': TimeBlock.objects.all()})


def runcodedatdown(response):
    if response.method == 'POST':
        user = response.user
        form = response.POST
        messagelist = []
        for time in TimeBlock.objects.all():
            id = 'q'+str(time.id)
            val = response.POST.get(id)
            bids = Bid.objects.filter(
                time=time.time, date=form.get('date'), is_down=True, is_dat=True).order_by('-price')
            quantity = float(val)
            if quantity > 0:
                sum = 0
                for bid in bids:
                    sum = sum+float(bid.volume)
                if sum == 0:
                    message = "Time Block:"+time.time+" Cleared Quantity: 0"
                    messagelist.append(message)
                    continue
                if sum < quantity:
                    message = "Time Block:"+time.time + \
                        " Cleared Quantity: "+str(sum)
                    messagelist.append(message)
                    # return redirect(reverse('runcodedatup'))
                    # return HttpResponse("The entered amount cannot be cleared.")
                else:
                    message = "Time Block:"+time.time + \
                        " Cleared Quantity: "+str(quantity)
                    messagelist.append(message)

                quantity = float(val)
                cleared = user.clearedreservedown_set.create(
                    time_block=time.time, date=form.get('date'), is_dat=True)
                for bid in bids:
                    if(float(bid.volume) < float(quantity)):
                        cleared.clearentitydown_set.create(
                            name=bid.seller.username, amount=bid.volume, price=bid.price)
                        quantity = quantity-float(bid.volume)
                    else:
                        cleared.clearentitydown_set.create(
                            name=bid.seller.username, amount=quantity, price=bid.price)
                        break
        return render(response, 'buyer/success.html', {"messages": messagelist})
    else:
        return render(response, "buyer/datform.html", {'timeblock': TimeBlock.objects.all()})


# Create your views here.
# data_fetch
# datafetch@123
