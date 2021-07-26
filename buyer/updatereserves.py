from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from .models import Reserve
margin = load_workbook('margin.xlsx')
for row in range(6, 102):
    for col in range(3, 15):
        cell = get_column_letter(col)
        Reserve.objects.create(
            time=margin['B'+str(row)], name=margin[cell+'5'], quantity=margin[cell+str(row)])
